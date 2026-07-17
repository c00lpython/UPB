import json
import os
import shutil
import openpyxl
from datetime import datetime
from PySide6.QtCore import QObject, Signal
from PySide6.QtWebEngineCore import QWebEngineProfile


class ProjectManager(QObject):
    """Управляет сохранением и загрузкой проектов"""
    
    project_loaded = Signal(dict)
    project_saved = Signal(str)
    
    def __init__(self):
        super().__init__()
        self.current_project = None
        self.current_profile = None
        self.projects_dir = os.path.join(os.getcwd(), "projects")
        self.latest_project_file = os.path.join(os.getcwd(), "latest_project.txt")
        
        if not os.path.exists(self.projects_dir):
            os.makedirs(self.projects_dir)

        self._cleanup_temp_profiles()
    
    def _cleanup_temp_profiles(self):
        """Очищает все временные профили браузера"""
        temp_pattern = "temp_profile_"
        cwd = os.getcwd()
        for item in os.listdir(cwd):
            item_path = os.path.join(cwd, item)
            if item.startswith(temp_pattern) and os.path.isdir(item_path):
                try:
                    shutil.rmtree(item_path, ignore_errors=True)
                except:
                    pass

    def save_latest_project(self, project_name: str):
        """Сохраняет последний открытый проект"""
        try:
            with open(self.latest_project_file, 'w', encoding='utf-8') as f:
                f.write(f"LatestProject = {project_name}\n")
        except:
            pass

    def get_latest_project(self) -> str:
        """Возвращает имя последнего открытого проекта"""
        try:
            if os.path.exists(self.latest_project_file):
                with open(self.latest_project_file, 'r', encoding='utf-8') as f:
                    for line in f:
                        if line.startswith("LatestProject = "):
                            project_name = line.split(" = ", 1)[1].strip()
                            project_path = os.path.join(self.projects_dir, project_name)
                            if os.path.exists(project_path):
                                return project_name
        except:
            pass
        return None

    def get_all_projects(self):
        projects = []
        if os.path.exists(self.projects_dir):
            for item in os.listdir(self.projects_dir):
                project_path = os.path.join(self.projects_dir, item)
                metadata_path = os.path.join(project_path, "metadata.json")
                if os.path.isdir(project_path) and os.path.exists(metadata_path):
                    try:
                        with open(metadata_path, 'r', encoding='utf-8') as f:
                            meta = json.load(f)
                        projects.append({
                            "name": item,
                            "created": meta.get("created", ""),
                            "modified": meta.get("modified", ""),
                            "variables_count": meta.get("variables_count", 0)
                        })
                    except:
                        projects.append({"name": item, "created": "", "modified": "", "variables_count": 0})
        return projects
    
    def _get_profile_for_project(self, project_name: str) -> QWebEngineProfile:
        """Создаёт или возвращает изолированный профиль для проекта"""
        profile_path = os.path.join(self.projects_dir, project_name, "profile")
        
        if not os.path.exists(profile_path):
            os.makedirs(profile_path)
        
        # Уникальное имя профиля = имя проекта (изоляция!)
        profile = QWebEngineProfile(project_name)
        profile.setPersistentStoragePath(profile_path)
        profile.setPersistentCookiesPolicy(
            QWebEngineProfile.PersistentCookiesPolicy.ForcePersistentCookies
        )
        profile.setHttpCacheType(QWebEngineProfile.HttpCacheType.DiskHttpCache)
        
        return profile
    
    def create_project(self, name: str):
        project_path = os.path.join(self.projects_dir, name)
        
        if os.path.exists(project_path):
            return False, "Project already exists"
        
        os.makedirs(project_path)
        os.makedirs(os.path.join(project_path, "parser"))
        os.makedirs(os.path.join(project_path, "profile"))  # Папка для изолированного профиля
        
        now = datetime.now().isoformat()
        
        metadata = {
            "name": name,
            "created": now,
            "modified": now,
            "version": "1.0",
            "variables_count": 0
        }
        with open(os.path.join(project_path, "metadata.json"), 'w', encoding='utf-8') as f:
            json.dump(metadata, f, indent=4, ensure_ascii=False)
        
        browser_data = {
            "tabs": [{"url": "https://google.com", "title": "Google"}],
            "current_tab": 0
        }
        with open(os.path.join(project_path, "browser_data.json"), 'w', encoding='utf-8') as f:
            json.dump(browser_data, f, indent=4, ensure_ascii=False)
        
        config_content = """[OUTPUT]
format = excel
excel_path = output.xlsx
telegram_bot_token = 
telegram_chat_id = 

[PARSER]
headless = false
timeout = 10
"""
        with open(os.path.join(project_path, "config.conf"), 'w', encoding='utf-8') as f:
            f.write(config_content)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Variables"
        ws.append(["Name", "XPath/CSS", "Type", "URL", "Sample Text"])
        wb.save(os.path.join(project_path, "variables.xlsx"))
        
        self.current_project = name
        self.current_profile = self._get_profile_for_project(name)
        self.save_latest_project(name)
        return True, {"path": project_path, "profile": self.current_profile}
    
    def load_project(self, name: str):
        project_path = os.path.join(self.projects_dir, name)
        
        if not os.path.exists(project_path):
            return None, "Project not found"
        
        with open(os.path.join(project_path, "metadata.json"), 'r', encoding='utf-8') as f:
            metadata = json.load(f)
        
        variables = []
        xlsx_path = os.path.join(project_path, "variables.xlsx")
        if os.path.exists(xlsx_path):
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    variables.append({
                        'name': row[0] or '',
                        'xpath': row[1] or '',
                        'type': row[2] or 'Static',
                        'url': row[3] or '',
                        'sample': row[4] or ''
                    })
        
        browser_data = {"tabs": [], "current_tab": 0}
        browser_path = os.path.join(project_path, "browser_data.json")
        if os.path.exists(browser_path):
            with open(browser_path, 'r', encoding='utf-8') as f:
                browser_data = json.load(f)
        
        config = {}
        config_path = os.path.join(project_path, "config.conf")
        if os.path.exists(config_path):
            with open(config_path, 'r', encoding='utf-8') as f:
                config_text = f.read()
                for line in config_text.split('\n'):
                    if '=' in line and not line.startswith('['):
                        key, value = line.split('=', 1)
                        config[key.strip()] = value.strip()
        
        self.current_project = name
        self.current_profile = self._get_profile_for_project(name)
        self.save_latest_project(name)
        return {
            "metadata": metadata,
            "variables": variables,
            "browser_data": browser_data,
            "config": config,
            "profile": self.current_profile
        }, "Loaded"
    
    def save_project(self, name: str, variables_data: list, browser_data: dict):
        project_path = os.path.join(self.projects_dir, name)
        
        if not os.path.exists(project_path):
            return False, "Project not found"
        
        now = datetime.now().isoformat()
        
        metadata_path = os.path.join(project_path, "metadata.json")
        with open(metadata_path, 'r', encoding='utf-8') as f:
            metadata = json.load(f)
        
        metadata["modified"] = now
        metadata["variables_count"] = len(variables_data)
        
        with open(metadata_path, 'w', encoding='utf-8') as f:
            json.dump(metadata, f, indent=4, ensure_ascii=False)
        
        with open(os.path.join(project_path, "browser_data.json"), 'w', encoding='utf-8') as f:
            json.dump(browser_data, f, indent=4, ensure_ascii=False)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Variables"
        ws.append(["Name", "XPath/CSS", "Type", "URL", "Sample Text"])
        
        for var in variables_data:
            ws.append([
                var.get('name', ''),
                var.get('xpath', ''),
                var.get('type', 'Static'),
                var.get('url', ''),
                var.get('sample', '')
            ])
        
        wb.save(os.path.join(project_path, "variables.xlsx"))
        self.save_latest_project(name)
        return True, "Project saved"
    
    def delete_project(self, name: str):
        project_path = os.path.join(self.projects_dir, name)
        
        if os.path.exists(project_path):
            shutil.rmtree(project_path)
            
            if self.current_project == name:
                self.current_project = None
                self.current_profile = None
            
            return True, "Project deleted"
        
        return False, "Project not found"
    
    def generate_parser(self, name: str, variables: list, config: dict):
        project_path = os.path.join(self.projects_dir, name)
        parser_dir = os.path.join(project_path, "parser")
        
        if not os.path.exists(parser_dir):
            os.makedirs(parser_dir)
        
        parser_code = self._generate_parser_code(name, variables, config)
        parser_file = os.path.join(parser_dir, f"{name}Parser.py")
        
        with open(parser_file, 'w', encoding='utf-8') as f:
            f.write(parser_code)
        
        return parser_file
    
    def _generate_parser_code(self, name: str, variables: list, config: dict):
        return f"""# {name}Parser.py
# Generated by UPB - Universal Parser Builder
# Date: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}

import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


class {name}Parser:
    def __init__(self, headless=False):
        options = webdriver.ChromeOptions()
        if headless:
            options.add_argument('--headless')
        self.driver = webdriver.Chrome(options=options)
        self.wait = WebDriverWait(self.driver, 10)
    
    def parse(self):
        variables = {{
{self._generate_variables_code(variables)}
        }}
        
        results = {{}}
        
        for name, xpath in variables.items():
            try:
                element = self.wait.until(
                    EC.presence_of_element_located((By.XPATH, xpath))
                )
                results[name] = element.text
                print(f"✓ {{name}}: {{results[name]}}")
            except Exception as e:
                print(f"✗ {{name}}: Error - {{e}}")
                results[name] = None
        
        return results
    
    def close(self):
        self.driver.quit()


def main():
    parser = {name}Parser(headless=False)
    
    try:
        parser.driver.get("https://example.com")
        time.sleep(2)
        
        results = parser.parse()
        
        df = pd.DataFrame([results])
        df.to_excel("output.xlsx", index=False)
        print("\\n✅ Results saved to output.xlsx")
        
    finally:
        parser.close()


if __name__ == "__main__":
    main()
"""
    
    def _generate_variables_code(self, variables: list) -> str:
        lines = []
        for var in variables:
            name = var.get('name', 'var')
            xpath = var.get('xpath', '')
            lines.append(f'            "{name}": "{xpath}",')
        return '\n'.join(lines)